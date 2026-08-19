import os
import json
import requests
import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(os.path.dirname(SCRIPT_DIR))
EXCEL_PATH = os.path.join(PROJECT_ROOT, 'InvesmateAPIs.xlsx')

LIB_DIR = os.path.join(PROJECT_ROOT, 'src', 'lib')
COMPANY_MASTER_OUT = os.path.join(LIB_DIR, 'companyMaster.json')
ETF_MASTER_OUT = os.path.join(LIB_DIR, 'etfMaster.json')
ISIN_MAP_OUT = os.path.join(LIB_DIR, 'isinMap.json')

def get_invesmate_token():
    print(f"Reading tokens from: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws_token = wb['Token']
    tokens = []
    for r in range(1, ws_token.max_row + 1):
        val = ws_token.cell(r, 2).value
        if val and str(val).startswith('eyJ'):
            tokens.append(str(val))
    
    for t in reversed(tokens):
        try:
            r = requests.get(
                'https://invesmateapis.cmots.com/api/CompanyMaster',
                headers={'Authorization': f'Bearer {t}'},
                timeout=8
            )
            if r.status_code == 200:
                print("Successfully authenticated with Invesmate API.")
                return t
        except Exception as err:
            print(f"Token test error: {err}")
    return None

def update_masters():
    token = get_invesmate_token()
    if not token:
        print("Error: No active Invesmate API token found.")
        return False

    headers = {'Authorization': f'Bearer {token}'}

    # 1. Fetch Company Master
    print("Fetching latest Company Master from Invesmate API...")
    r_comp = requests.get('https://invesmateapis.cmots.com/api/CompanyMaster', headers=headers, timeout=30)
    comp_json = r_comp.json()
    comp_raw_list = comp_json.get('data', [])
    print(f"Received {len(comp_raw_list)} companies.")

    companies = []
    nse_index = {}
    bse_index = {}
    isin_index = {}
    isin_to_nse = {}

    for item in comp_raw_list:
        comp_name = str(item.get('companyname') or '').strip()
        sector = str(item.get('sectorname') or '').strip()
        mcap_raw = str(item.get('mcaptype') or '').strip()
        
        # Standardize market cap category
        mcap_type = 'Mid'
        if 'large' in mcap_raw.lower():
            mcap_type = 'Large'
        elif 'small' in mcap_raw.lower():
            mcap_type = 'Small'

        mcap_val = float(item.get('mcap') or 0.0)
        industry = str(item.get('industryname') or '').strip()
        isin = str(item.get('isin') or '').strip().upper()
        bse_group = str(item.get('bsegroup') or '').strip()
        short_name = str(item.get('companyshortname') or '').strip()
        nse_status = str(item.get('NSEStatus') or '').strip()
        bse_status = str(item.get('BSEStatus') or '').strip()

        # Tuple structure matching sectorMap.ts:
        # [0]name, [1]sector, [2]mcaptype, [3]mcapVal, [4]industry, [5]isin, [6]bsegroup, [7]shortname, [8]nseStatus, [9]bseStatus
        tuple_row = [
            comp_name,
            sector,
            mcap_type,
            round(mcap_val, 2),
            industry,
            isin,
            bse_group,
            short_name,
            nse_status,
            bse_status
        ]

        idx = len(companies)
        companies.append(tuple_row)

        nse_sym = str(item.get('nsesymbol') or '').strip().upper()
        bse_code = str(item.get('bsecode') or '').strip()

        if nse_sym:
            nse_index[nse_sym] = idx
            if isin:
                isin_to_nse[isin] = nse_sym

        if bse_code:
            bse_index[bse_code] = idx

        if isin:
            isin_index[isin] = idx

    company_master_dict = {
        'nse': nse_index,
        'bse': bse_index,
        'isin': isin_index,
        'companies': companies
    }

    with open(COMPANY_MASTER_OUT, 'w', encoding='utf-8') as f:
        json.dump(company_master_dict, f, separators=(',', ':'))
    print(f"Updated {COMPANY_MASTER_OUT} with {len(companies)} companies.")

    # 2. Fetch ETF Master
    print("\nFetching latest ETF Master from Invesmate API...")
    r_etf = requests.get('https://invesmateapis.cmots.com/api/ETFMaster', headers=headers, timeout=30)
    etf_json = r_etf.json()
    etf_raw_list = etf_json.get('data', [])
    print(f"Received {len(etf_raw_list)} ETFs.")

    etfs = []
    etf_isin_index = {}
    ticker_index = {}

    for item in etf_raw_list:
        etf_name = str(item.get('ETFName') or '').strip()
        category = str(item.get('ETFCategory') or '').strip()
        amc_name = str(item.get('AMCName') or '').strip()
        isin = str(item.get('ISIN') or '').strip().upper()

        if not isin or isin == 'UNDEFINED':
            continue

        tuple_row = [etf_name, category, amc_name]
        idx = len(etfs)
        etfs.append(tuple_row)
        etf_isin_index[isin] = idx

        # Common Ticker Mapping
        name_lower = etf_name.lower()
        if 'gold bees' in name_lower or 'goldbees' in name_lower:
            ticker_index['GOLDBEES'] = idx
        elif 'silver bees' in name_lower or 'silverbees' in name_lower:
            ticker_index['SILVERBEES'] = idx
        elif 'junior bees' in name_lower or 'juniorbees' in name_lower or 'next 50 junior' in name_lower:
            ticker_index['JUNIORBEES'] = idx
        elif 'liquid bees' in name_lower or 'liquidbees' in name_lower or '1d rate liquid' in name_lower:
            ticker_index['LIQUIDBEES'] = idx
        elif 'nifty 50 bees' in name_lower or 'nifty bees' in name_lower:
            ticker_index['NIFTYBEES'] = idx
        elif 'bank bees' in name_lower or 'bankbees' in name_lower or 'nifty bank bees' in name_lower:
            ticker_index['BANKBEES'] = idx
        elif 'it bees' in name_lower or 'itbees' in name_lower:
            ticker_index['ITBEES'] = idx
        elif 'hang seng bees' in name_lower or 'hangsengbees' in name_lower:
            ticker_index['HANGSENGBEES'] = idx
        elif 'pharma bees' in name_lower or 'pharmabees' in name_lower:
            ticker_index['PHARMABEES'] = idx
        elif 'infra bees' in name_lower or 'infrabees' in name_lower:
            ticker_index['INFRABEES'] = idx
        elif 'consumption bees' in name_lower or 'consumptionbees' in name_lower:
            ticker_index['CONSUMPTIONBEES'] = idx
        elif 'monifty50' in name_lower:
            ticker_index['MONIFTY50'] = idx
        elif 'nifty100' in name_lower:
            ticker_index['NIFTY100'] = idx
        elif 'midcap100' in name_lower:
            ticker_index['MIDCAP100'] = idx
        elif 'nasdaq 100' in name_lower or 'nasdaq100' in name_lower:
            ticker_index['MONASDAQ100'] = idx

    etf_master_dict = {
        'etfs': etfs,
        'isin': etf_isin_index,
        'ticker': ticker_index
    }

    with open(ETF_MASTER_OUT, 'w', encoding='utf-8') as f:
        json.dump(etf_master_dict, f, separators=(',', ':'))
    print(f"Updated {ETF_MASTER_OUT} with {len(etfs)} ETFs.")

    # 3. Update isinMap.json
    with open(ISIN_MAP_OUT, 'w', encoding='utf-8') as f:
        json.dump(isin_to_nse, f, separators=(',', ':'))
    print(f"Updated {ISIN_MAP_OUT} with {len(isin_to_nse)} ISIN-to-NSE mappings.")

    print("\nMaster JSON datasets updated successfully!")
    return True

if __name__ == '__main__':
    update_masters()
